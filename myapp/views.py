import csv
from django.db.models import Sum, Max
from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse
from datetime import datetime
from rest_framework.response import Response
from rest_framework.decorators import api_view
import json
from io import TextIOWrapper
from .forms import CSVUploadForm, InvoiceUpdateForm, InvoiceCreateForm, InvoiceForm, InvoiceDetailForm
from .models import Invoice, Store, CustomerGroup, Customer, ProductCategory, Product, InvoiceDetail



# Trang chủ blog cây cảnh
def home(request):
    return render(request, 'myapp/home.html')


def invoice_list(request):
    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['file']
            decoded_file = TextIOWrapper(csv_file.file, encoding='utf-8', errors='replace')
            reader = csv.reader(decoded_file, delimiter=',')

            next(reader)

            stores_dict = {}
            customer_groups_dict = {}
            customers_dict = {}
            product_categories_dict = {}
            products_dict = {}
            invoices_dict = {}
            invoice_details = []

            for row in reader:
                try:
                    if row[1] not in stores_dict:
                        store = Store(ma_cua_hang=row[1], doanh_nghiep=row[0], dia_chi=row[2])
                        stores_dict[row[1]] = store

                    if row[6] not in customer_groups_dict:
                        customer_group = CustomerGroup(ma_nhom_kh=row[6], thong_tin_nhom_kh=row[7])
                        customer_groups_dict[row[6]] = customer_group

                    if row[8] not in customers_dict:
                        customer = Customer(ma_kh=row[8], ma_nhom_kh=customer_groups_dict[row[6]])
                        customers_dict[row[8]] = customer

                    if row[9] not in product_categories_dict:
                        product_category = ProductCategory(ma_nhom_hang=row[9], nhom_hang=row[9])
                        product_categories_dict[row[9]] = product_category

                    if row[11] not in products_dict:
                        product = Product(
                            ma_hang=row[11],
                            ma_nhom_hang=product_categories_dict[row[9]],
                            mat_hang=row[12],
                            dvt=row[13],
                            don_gia=float(row[15])
                        )
                        products_dict[row[11]] = product

                    if row[5] not in invoices_dict:
                        invoice = Invoice(
                            ma_hoa_don=row[5],
                            ma_cua_hang=stores_dict[row[1]],
                            ma_kh=customers_dict[row[8]],
                            nam=int(row[3]),
                            thang=int(row[4])
                        )
                        invoices_dict[row[5]] = invoice

                    invoice_detail = InvoiceDetail(
                        invoice=invoices_dict[row[5]],
                        ma_hang=products_dict[row[11]],
                        sl_ban=int(row[14]),
                        tam_tinh=float(row[15])
                    )
                    invoice_details.append(invoice_detail)

                except Exception as e:
                    invoices = Invoice.objects.all()
                    messages.error(request, "Import invoice failed, data not match the struct")
                    return render(request, 'myapp/invoice_list.html', {
                        'form': form,
                        'invoices': invoices
                    })

            Store.objects.bulk_create(stores_dict.values(), ignore_conflicts=True)
            CustomerGroup.objects.bulk_create(customer_groups_dict.values(), ignore_conflicts=True)
            Customer.objects.bulk_create(customers_dict.values(), ignore_conflicts=True)
            ProductCategory.objects.bulk_create(product_categories_dict.values(), ignore_conflicts=True)
            Product.objects.bulk_create(products_dict.values(), ignore_conflicts=True)
            Invoice.objects.bulk_create(invoices_dict.values(), ignore_conflicts=True)
            InvoiceDetail.objects.bulk_create(invoice_details, ignore_conflicts=True)
            messages.success(request, "Import invoice successed")

    else:
        form = CSVUploadForm()

    invoices = Invoice.objects.annotate(tong_gia=Sum('invoicedetail__tam_tinh')).order_by("-ma_hoa_don")
    paginator = Paginator(invoices, 10)  # Số lượng hóa đơn mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'myapp/invoice_list.html', {
        'form': form,
        'page_obj': page_obj,
    })

def invoice_detail(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)
    
    # Tính tổng giá từ InvoiceDetail
    total_price = invoice.invoicedetail_set.aggregate(Sum('tam_tinh'))['tam_tinh__sum'] or 0

    if request.method == 'POST':
        form = InvoiceUpdateForm(request.POST, instance=invoice)
        if form.is_valid():
            form.save()
            messages.success(request, "Invoice updated successfully.")
            return redirect(reverse('invoice_detail', kwargs={'pk': invoice.pk}))
    else:
        form = InvoiceUpdateForm(instance=invoice)

    return render(request, 'myapp/invoice_detail.html', {
        'invoice': invoice,
        'total_price': total_price,
        'form': form,
    })

def delete_invoice_detail(request, detail_id):
    # Lấy đối tượng InvoiceDetail dựa trên ID
    detail = get_object_or_404(InvoiceDetail, pk=detail_id)
    invoice = detail.invoice  # Lấy đối tượng hóa đơn tương ứng
    invoice_id = invoice.ma_hoa_don  # Lưu lại ID của hóa đơn trước khi xóa

    # Xóa đối tượng InvoiceDetail
    detail.delete()

    # Kiểm tra số lượng chi tiết hóa đơn còn lại
    if invoice.invoicedetail_set.count() == 0:
        # Nếu không còn chi tiết nào, xóa hóa đơn
        invoice.delete()
        messages.success(request, "Invoice and product removed successfully.")
        return redirect('invoice_list')
    else:
        messages.success(request, "Product removed from invoice successfully.")
    
    # Chuyển hướng về trang chi tiết hóa đơn
    return redirect(reverse('invoice_detail', kwargs={'pk': invoice_id}))

def update_invoice_detail(request, id):
    # Lấy đối tượng InvoiceDetail dựa trên ID
    invoice_detail = get_object_or_404(InvoiceDetail, id=id)
    
    if request.method == 'POST':
        # Lấy số lượng mới từ form
        new_quantity = request.POST.get('quantity')
        
        if new_quantity.isdigit() and int(new_quantity) > 0:
            # Cập nhật số lượng
            invoice_detail.sl_ban = int(new_quantity)
            invoice_detail.tam_tinh = invoice_detail.ma_hang.don_gia * int(new_quantity)
            invoice_detail.save()
            messages.success(request, 'Quantity updated successfully!')
        else:
            messages.error(request, 'Invalid quantity. Please enter a positive number.')
    
    # Chuyển hướng trở lại trang chi tiết hóa đơn
    # return redirect('invoice_detail', id=invoice_detail.invoice.ma_hoa_don)
    return redirect(reverse('invoice_detail', kwargs={'pk': invoice_detail.invoice.ma_hoa_don}))

def generate_customer_code():
    # Tìm mã khách hàng lớn nhất trong database
    latest_customer = Customer.objects.aggregate(max_code=Max('ma_kh'))
    max_code = latest_customer['max_code']
    
    if max_code:
        # Lấy phần số từ mã khách hàng hiện tại, giả sử mã khách hàng là 'CUS0000296'
        latest_number = int(max_code[3:])  # Lấy từ vị trí thứ 3 trở đi (bỏ 'CUS')
        new_number = latest_number + 1
    else:
        # Nếu chưa có mã khách hàng nào trong database, bắt đầu từ số 1
        new_number = 1

    # Tạo mã khách hàng mới theo định dạng 'CUS0000XXXX' với phần số tăng lên
    new_code = f"CUS{new_number:07}"  # 07 để điền đủ 7 chữ số, ví dụ: CUS0000031

    return new_code


@api_view(['POST'])
def create_invoice1(request):
    data = json.loads(request.body)
    ma_cua_hang = data.get('ma_cua_hang')
    ma_nhom_kh = data.get('ma_kh')
    nam = datetime.now().year
    thang = datetime.now().month
    product_ids = data.get('product_ids', [])
    quantities = data.get('quantities', [])
    totalprice = data.get('tam_tinh', [])
    

    store = get_object_or_404(Store, ma_cua_hang=ma_cua_hang)
    customer_id = generate_customer_code()
    customer_group = CustomerGroup.objects.get(ma_nhom_kh=ma_nhom_kh)

    # Tạo mã khách hàng và lưu
    new_customer = Customer(ma_kh=customer_id, ma_nhom_kh=customer_group)
    new_customer.save()

    # Lấy mã hóa đơn cuối cùng và tạo mã hóa đơn mới
    last_invoice = Invoice.objects.aggregate(Max('ma_hoa_don'))
    last_invoice_code = last_invoice['ma_hoa_don__max']
    if last_invoice_code:
        new_invoice_number = int(last_invoice_code[1:]) + 1
        new_invoice_code = f"B{new_invoice_number:09d}"
    else:
        new_invoice_code = "B000000001"

    # Tạo hóa đơn mới
    invoice = Invoice(ma_hoa_don=new_invoice_code, ma_cua_hang=store, ma_kh=new_customer, nam=nam, thang=thang, )
    invoice.save()

    # Tạo các chi tiết hóa đơn
    invoice_details = []
    for product_id, quantity in zip(product_ids, quantities):
        if int(quantity) > 0:
            invoice_details.append(InvoiceDetail(invoice=invoice, ma_hang_id=product_id, sl_ban=quantity,tam_tinh=totalprice,))

    if invoice_details:
        InvoiceDetail.objects.bulk_create(invoice_details)

    # Lấy danh sách các hóa đơn đã được tạo
    invoices = Invoice.objects.order_by('ma_hoa_don').values(
        'ma_hoa_don', 'ma_cua_hang__ma_cua_hang', 'ma_kh__ma_kh', 'nam', 'thang',
    )

    # Chuẩn bị dữ liệu trả về
    response_data = {
        'success': True,
        'invoices': list(invoices),  # Trả về danh sách hóa đơn mới nhất
        'redirect_url': '/invoices/',  # Đường dẫn đến trang danh sách hóa đơn nếu cần
    }

    return JsonResponse(response_data)

    


def create_invoice(request):
    if request.method == 'POST':
        # Lấy dữ liệu từ JSON
        data = json.loads(request.body)

        ma_cua_hang = data.get('ma_cua_hang')
        ma_kh = data.get('ma_kh')
        nam = datetime.now().year
        thang = datetime.now().month
        product_ids = data.get('product_ids', [])
        quantities = data.get('quantities', [])

        store = get_object_or_404(Store, ma_cua_hang=ma_cua_hang)
        customer = get_object_or_404(Customer, ma_kh=ma_kh)

        # Tạo Invoice
        invoice = Invoice(ma_cua_hang=store, ma_kh=customer, nam=nam, thang=thang)
        invoice.save()

        # Tạo danh sách InvoiceDetail
        invoice_details = []
        for product_id, quantity in zip(product_ids, quantities):
            # Chỉ tạo InvoiceDetail nếu quantity hợp lệ
            if quantity > 0:  # Kiểm tra số lượng phải lớn hơn 0
                invoice_details.append(InvoiceDetail(invoice=invoice, ma_hang_id=product_id, sl_ban=quantity))

        # Sử dụng bulk_create để lưu tất cả InvoiceDetail trong một lần
        if invoice_details:
            InvoiceDetail.objects.bulk_create(invoice_details)

        # Trả về phản hồi JSON
        return JsonResponse({'message': 'Invoice created successfully!', 'redirect_url': 'invoice_success'})

    # Nếu không phải POST, trả về form như trước
    products = Product.objects.all()
    customer_groups = CustomerGroup.objects.all()
    stores = Store.objects.all()

    context = {
        'products': products,
        'customer_groups': customer_groups,
        'stores': stores,
    }
    return render(request, 'myapp/create_invoice.html', context)





from django.shortcuts import render, redirect
from django.http import JsonResponse
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from .models import Store, CustomerGroup, Customer, ProductCategory, Product, Invoice, InvoiceDetail
from .forms import UploadFileForm
from openpyxl import load_workbook

@csrf_exempt
def import_excel_view(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Đọc file Excel
                wb = load_workbook(filename=file)
                ws = wb.active

                # Lấy dữ liệu từ Excel và lưu vào database
                data_rows = list(ws.iter_rows(values_only=True))[1:]  # Bỏ qua dòng tiêu đề
                print(data_rows)
                save_data_to_db(data_rows)

                messages.success(request, "Dữ liệu từ file Excel đã được nhập thành công.")
            except Exception as e:
                print(e)
                messages.error(request, "Lỗi nhập dữ liệu từ file Excel.")
        else:
            messages.error(request, "Form không hợp lệ.")
    
    return redirect('invoice_list')

def get_or_create_instance(model, filters, defaults=None):
    instance, _ = model.objects.get_or_create(**filters, defaults=defaults)
    return instance

def save_data_to_db(data_rows):
    stores_dict = {}
    customer_groups_dict = {}
    customers_dict = {}
    product_categories_dict = {}
    products_dict = {}
    invoices_dict = {}
    invoice_details = []

    for row in data_rows:
        try:
            if row[1] not in stores_dict:
                store = Store(ma_cua_hang=row[1], doanh_nghiep=row[0], dia_chi=row[2])
                stores_dict[row[1]] = store

            if row[6] not in customer_groups_dict:
                customer_group = CustomerGroup(ma_nhom_kh=row[6], thong_tin_nhom_kh=row[7])
                customer_groups_dict[row[6]] = customer_group

            if row[8] not in customers_dict:
                customer = Customer(ma_kh=row[8], ma_nhom_kh=customer_groups_dict[row[6]])
                customers_dict[row[8]] = customer

            if row[9] not in product_categories_dict:
                product_category = ProductCategory(ma_nhom_hang=row[9], nhom_hang=row[9])
                product_categories_dict[row[9]] = product_category

            if row[11] not in products_dict:
                product = Product(
                    ma_hang=row[11],
                    ma_nhom_hang=product_categories_dict[row[9]],
                    mat_hang=row[12],
                    dvt=row[13],
                    don_gia=float(row[15])
                )
                products_dict[row[11]] = product

            if row[5] not in invoices_dict:
                invoice = Invoice(
                    ma_hoa_don=row[5],
                    ma_cua_hang=stores_dict[row[1]],
                    ma_kh=customers_dict[row[8]],
                    nam=int(row[3]),
                    thang=int(row[4])
                )
                invoices_dict[row[5]] = invoice

            invoice_detail = InvoiceDetail(
                invoice=invoices_dict[row[5]],
                ma_hang=products_dict[row[11]],
                sl_ban=int(row[14]),
                tam_tinh=float(row[15])
            )
            invoice_details.append(invoice_detail)

        except Exception as e:
            print("ABCAS")
            print("ERROR RIGHT HErre", e)

    Store.objects.bulk_create(stores_dict.values(), ignore_conflicts=True)
    CustomerGroup.objects.bulk_create(customer_groups_dict.values(), ignore_conflicts=True)
    Customer.objects.bulk_create(customers_dict.values(), ignore_conflicts=True)
    ProductCategory.objects.bulk_create(product_categories_dict.values(), ignore_conflicts=True)
    Product.objects.bulk_create(products_dict.values(), ignore_conflicts=True)
    Invoice.objects.bulk_create(invoices_dict.values(), ignore_conflicts=True)
    InvoiceDetail.objects.bulk_create(invoice_details, ignore_conflicts=True)
    print("done")




from django.shortcuts import render
from .models import Invoice, InvoiceDetail, Store, Product, CustomerGroup, ProductCategory
from django.core.paginator import Paginator

def export_invoice_data(request):
    # Truy vấn tất cả các hóa đơn và chi tiết hóa đơn
    invoices = Invoice.objects.all()
    
    # Paginate kết quả nếu cần
    paginator = Paginator(invoices, 10)  # Hiển thị 10 hóa đơn mỗi trang
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Tạo danh sách dữ liệu để hiển thị trong bảng
    invoice_data = []
    for invoice in page_obj:
        invoice_details = InvoiceDetail.objects.filter(invoice=invoice)
        for detail in invoice_details:
            invoice_data.append({
                'ma_cua_hang': invoice.ma_cua_hang.ma_cua_hang,
                'nam': invoice.nam,
                'thang': invoice.thang,
                'ma_hoa_don': invoice.ma_hoa_don,
                'thong_tin_nhom_kh': invoice.ma_kh.ma_nhom_kh.thong_tin_nhom_kh if invoice.ma_kh else '',
                'nhom_hang': detail.ma_hang.ma_nhom_hang.nhom_hang if detail.ma_hang else '',
                'mat_hang': detail.ma_hang.mat_hang if detail.ma_hang else '',
                'dvt': detail.ma_hang.dvt if detail.ma_hang else '',
                'sl_ban': detail.sl_ban,
                'don_gia': detail.ma_hang.don_gia if detail.ma_hang else 0,
                'tam_tinh': detail.tam_tinh,
            })

    return render(request, 'myapp/export_invoice_data.html', {
        'page_obj': page_obj,
        'invoice_data': invoice_data
    })


from django.shortcuts import render
from .models import Invoice, InvoiceDetail
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Invoice, InvoiceDetail, CustomerGroup, Product
from django.db.models import Sum

def chart_view(request):
    # Tổng số lượng sản phẩm bán được theo mã hàng
    product_sales = InvoiceDetail.objects.values('ma_hang__mat_hang').annotate(total_sales=Sum('sl_ban'))
    product_labels = [item['ma_hang__mat_hang'] for item in product_sales]
    product_sales_data = [item['total_sales'] for item in product_sales]

    # Phân bố nhóm khách hàng
    customer_groups = CustomerGroup.objects.values('thong_tin_nhom_kh').annotate(count=Sum('customer__ma_kh'))
    customer_group_labels = [item['thong_tin_nhom_kh'] for item in customer_groups]
    customer_group_data = [item['count'] for item in customer_groups]

    # Doanh thu theo tháng trong năm 2020
    monthly_revenue = InvoiceDetail.objects.filter(invoice__nam=2020).values('invoice__thang').annotate(revenue=Sum('tam_tinh'))
    monthly_revenue_labels = [f"Tháng {item['invoice__thang']}" for item in monthly_revenue]
    monthly_revenue_data = [item['revenue'] for item in monthly_revenue]

    context = {
        'product_labels': product_labels,
        'product_sales_data': product_sales_data,
        'customer_group_labels': customer_group_labels,
        'customer_group_data': customer_group_data,
        'monthly_revenue_labels': monthly_revenue_labels,
        'monthly_revenue_data': monthly_revenue_data,
    }
    return render(request, 'myapp/chart.html', context)